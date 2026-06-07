"""Excel文档解析器"""

from typing import Any, Dict, List

from loguru import logger

from app.rag.parsers.base import BaseParser


class ExcelParser(BaseParser):
    """Excel文档解析器，使用openpyxl解析.xlsx文件"""

    # 每个chunk包含的最大行数
    ROWS_PER_CHUNK: int = 20

    def parse(self, file_path: str) -> List[Dict[str, Any]]:
        """解析Excel文档，按sheet分割，每若干行作为一个chunk

        Args:
            file_path: Excel文件路径

        Returns:
            解析结果列表:
                - content: 表格文本内容
                - metadata: 包含sheet_name、row_range等信息

        Raises:
            FileNotFoundError: 文件不存在
            Exception: Excel解析失败
        """
        from openpyxl import load_workbook

        results: List[Dict[str, Any]] = []

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            logger.info(f"开始解析Excel文件: {file_path}，共 {len(wb.sheetnames)} 个Sheet")

            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    sheet_results = self._parse_sheet(ws, sheet_name, file_path)
                    results.extend(sheet_results)
                except Exception as e:
                    logger.warning(f"Sheet '{sheet_name}' 解析失败: {e}")
                    continue

            wb.close()
            logger.info(f"Excel解析完成，共提取 {len(results)} 个内容块")

        except FileNotFoundError:
            raise FileNotFoundError(f"文件不存在: {file_path}")
        except Exception as e:
            logger.error(f"Excel解析失败: {e}")
            raise Exception(f"Excel解析失败: {e}")

        return results

    def _parse_sheet(
        self, worksheet, sheet_name: str, file_path: str
    ) -> List[Dict[str, Any]]:
        """解析单个Sheet

        Args:
            worksheet: openpyxl worksheet对象
            sheet_name: Sheet名称
            file_path: 文件路径

        Returns:
            该Sheet的解析结果列表
        """
        results: List[Dict[str, Any]] = []

        # 读取所有行数据
        all_rows: List[List[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            # 将None替换为空字符串，统一转为字符串
            row_data = [str(cell).strip() if cell is not None else "" for cell in row]
            # 跳过全空行
            if any(cell for cell in row_data):
                all_rows.append(row_data)

        if not all_rows:
            return results

        # 第一行作为表头
        header = all_rows[0]
        header_text = " | ".join(header)

        # 数据行按ROWS_PER_CHUNK分组
        data_rows = all_rows[1:]
        for chunk_start in range(0, len(data_rows), self.ROWS_PER_CHUNK):
            chunk_end = min(chunk_start + self.ROWS_PER_CHUNK, len(data_rows))
            chunk_rows = data_rows[chunk_start:chunk_end]

            # 构建文本：表头 + 数据行
            text_parts = [f"[表头] {header_text}"]
            for row in chunk_rows:
                row_text = " | ".join(row)
                text_parts.append(row_text)

            chunk_text = "\n".join(text_parts)
            cleaned = self.clean_text(chunk_text)

            if cleaned:
                results.append(
                    {
                        "content": cleaned,
                        "metadata": {
                            "sheet_name": sheet_name,
                            "row_start": chunk_start + 2,  # +2因为从第2行开始（1-indexed + 跳过表头）
                            "row_end": chunk_end + 1,
                            "total_rows": len(all_rows),
                            "source": file_path,
                            "type": "excel_table",
                        },
                    }
                )

        return results
